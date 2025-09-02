import os
from openpyxl import load_workbook

def get_client_data_xlsx(file_name, first_name, last_name):
    base_dir = os.path.dirname(__file__)              # adresář, kde je skript
    file_path = os.path.join(base_dir, file_name)     # úplná cesta k souboru
    
    wb = load_workbook(filename=file_path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        jmeno, prijmeni = row[0], row[1]
        if jmeno and prijmeni and jmeno.lower() == first_name.lower() and prijmeni.lower() == last_name.lower():
            return dict(zip(headers, row))

    return None

# Příklad volání:
info = get_client_data_xlsx("klienti.xlsx", "Jan", "Novák")
print(info)