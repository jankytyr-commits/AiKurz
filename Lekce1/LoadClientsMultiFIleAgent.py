import os
from openpyxl import Workbook, load_workbook
from collections import OrderedDict

# --- Krok 1: Pomocná funkce pro vytvoření testovacích souborů ---
def setup_dummy_excel_files():
    """
    Vytvoří dva testovací Excel soubory s 10 záznamy celkem
    v adresáři, kde se nachází spuštěný skript.
    """
    # Zajištění, že cesta je relativní k umístění skriptu
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    files_to_create = {
        "klienti_praha.xlsx": [
            ("Jan", "Novák", "jan.novak@email.cz", "Praha", 34),
            ("Jana", "Nováková", "jana.novakova@email.cz", "Praha", 31),
            ("Martin", "Dlouhý", "martin.dlouhy@email.cz", "Praha", 45),
            ("Veronika", "Bílá", "veronika.bila@email.cz", "Praha", 22),
            ("Lukáš", "Marek", "lukas.marek@email.cz", "Praha", 39),
        ],
        "klienti_brno.xlsx": [
            ("Petra", "Svobodová", "petra.s@email.cz", "Brno", 28),
            ("Pavel", "Dvořák", "pavel.d@email.cz", "Brno", 52),
            ("Alena", "Veselá", "alena.vesela@email.cz", "Brno", 41),
            ("Tomáš", "Horák", "tomas.horak@email.cz", "Brno", 36),
            ("Eva", "Kučerová", "eva.kucerova@email.cz", "Brno", 29),
        ]
    }
    
    headers = ["Jméno", "Příjmení", "Email", "Město", "Věk"]
    created_paths = []

    for file_name, data_rows in files_to_create.items():
        file_path = os.path.join(base_dir, file_name)
        created_paths.append(file_path)

        if os.path.exists(file_path):
            os.remove(file_path)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Klientů"
        ws.append(headers)
        for row in data_rows:
            ws.append(row)
        wb.save(file_path)
        print(f"Testovací soubor '{file_path}' byl vytvořen.")
    
    return created_paths

# --- Krok 2: Nový Agent pro více souborů ---
class MultiExcelAgent:
    """
    Agent pro správu dat z více Excel souborů najednou.
    """
    def __init__(self, file_paths: list):
        """
        Inicializace agenta se seznamem cest k souborům.
        """
        self.data_sources = OrderedDict() # Uloží data pro každý soubor zvlášť
        self._load_files(file_paths)

    def _load_files(self, file_paths: list):
        """Načte data ze všech zadaných souborů do paměti."""
        for path in file_paths:
            if not os.path.exists(path):
                print(f"Varování: Soubor '{path}' nebyl nalezen a bude přeskočen.")
                continue
            
            wb = load_workbook(filename=path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            file_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                record = OrderedDict(zip(headers, row))
                record['_source_file'] = path # Klíčové: uložíme si původ souboru!
                file_data.append(record)
            
            self.data_sources[path] = {'headers': headers, 'data': file_data}
            print(f"Agent: Data z '{path}' byla úspěšně načtena.")

    def _uloz_zmeny(self, file_path: str):
        """Zapíše aktuální stav dat pro daný soubor zpět na disk."""
        if file_path not in self.data_sources:
            print(f"Chyba: Nelze uložit. Soubor '{file_path}' není spravován tímto agentem.")
            return

        source = self.data_sources[file_path]
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Klientů"
        ws.append(source['headers'])
        
        for record in source['data']:
            # Vytvoříme kopii, abychom mohli smazat interní klíč '_source_file'
            record_to_save = record.copy()
            del record_to_save['_source_file']
            ws.append(list(record_to_save.values()))
            
        wb.save(filename=file_path)
        # print(f"Agent: Změny byly uloženy do souboru '{file_path}'.")

    def _get_all_records(self):
        """Interní generátor pro procházení všech záznamů ze všech souborů."""
        for source in self.data_sources.values():
            for record in source['data']:
                yield record

    # --- CRUD Metody ---
    def najdi_vsechny_zaznamy(self, **kwargs):
        """Najde všechny záznamy napříč všemi soubory, které odpovídají kritériím."""
        search_criteria = {k: str(v).lower() for k, v in kwargs.items()}
        results = []
        for record in self._get_all_records():
            match = all(str(record.get(key)).lower() == value for key, value in search_criteria.items())
            if match:
                results.append(record)
        return results

    def najdi_zaznam(self, **kwargs):
        """Najde první záznam napříč všemi soubory."""
        results = self.najdi_vsechny_zaznamy(**kwargs)
        return results[0] if results else None

    def pridej_zaznam(self, novy_zaznam: dict, file_path: str):
        """Přidá nový záznam do specifikovaného souboru."""
        if file_path not in self.data_sources:
            print(f"Chyba: Soubor '{file_path}' není spravován. Záznam nelze přidat.")
            return False
        
        source = self.data_sources[file_path]
        record = OrderedDict((h, novy_zaznam.get(h)) for h in source['headers'])
        record['_source_file'] = file_path
        
        source['data'].append(record)
        self._uloz_zmeny(file_path)
        return True

    def uprav_zaznam(self, zmeny: dict, kriteria: dict):
        """Najde první záznam dle kritérií (napříč soubory) a aktualizuje ho."""
        record_to_update = self.najdi_zaznam(**kriteria)
        if not record_to_update:
            return False
        
        source_file = record_to_update['_source_file']
        record_to_update.update(zmeny)
        self._uloz_zmeny(source_file)
        return True

    def smaz_zaznam(self, **kwargs):
        """Najde první záznam dle kritérií (napříč soubory) a smaže ho."""
        record_to_delete = self.najdi_zaznam(**kwargs)
        if not record_to_delete:
            return False
            
        source_file = record_to_delete['_source_file']
        # Odstraníme záznam ze seznamu v paměti
        self.data_sources[source_file]['data'].remove(record_to_delete)
        self._uloz_zmeny(source_file)
        return True

# --- Krok 3: Příklad použití multi-file agenta ---
if __name__ == "__main__":
    # 1. Vytvoříme testovací soubory a získáme jejich cesty
    file_paths = setup_dummy_excel_files()
    
    print("\n" + "="*40)
    print("SPOUŠTÍM AGENTA PRO VÍCE SOUBORŮ")
    print("="*40 + "\n")

    # 2. Vytvoříme agenta, který spravuje oba soubory
    agent = MultiExcelAgent(file_paths=file_paths)

    print("\n--- 1. Hledání všech záznamů z Brna ---")
    brno_records = agent.najdi_vsechny_zaznamy(Město="Brno")
    print(f"Nalezeno {len(brno_records)} záznamů z Brna:")
    for rec in brno_records:
        print(f" > {rec['Jméno']} {rec['Příjmení']}")


    print("\n--- 2. Přidání záznamu do konkrétního souboru ---")
    praha_file_path = next(path for path in file_paths if 'praha' in path)
    print(f"Přidávám 'Petr Král' do souboru '{praha_file_path}'")
    novy_klient = {"Jméno": "Petr", "Příjmení": "Král", "Email": "p.kral@email.com", "Město": "Praha", "Věk": 55}
    agent.pridej_zaznam(novy_klient, file_path=praha_file_path)
    print("Kontrola - hledám Petra Krále:")
    karel = agent.najdi_zaznam(Příjmení="Král")
    print(f"Nalezen: {dict(karel)}")

    print("\n--- 3. Úprava záznamu (agent sám pozná soubor) ---")
    print("Měním věk Evy Kučerové na 30 let.")
    agent.uprav_zaznam(zmeny={"Věk": 30}, kriteria={"Email": "eva.kucerova@email.cz"})
    print("Kontrola - hledám Evu Kučerovou:")
    jana = agent.najdi_zaznam(Příjmení="Kučerová")
    print(f"Nalezen: {dict(jana)}")

    print("\n--- 4. Smazání záznamu (agent sám pozná soubor) ---")
    print("Mažu Pavla Dvořáka.")
    uspech = agent.smaz_zaznam(Příjmení="Dvořák")
    print(f"Operace smazání byla {'úspěšná' if uspech else 'neúspěšná'}.")
    print("Kontrola - hledám Pavla Dvořáka:")
    pavel = agent.najdi_zaznam(Příjmení="Dvořák")
    print(f"Výsledek hledání: {pavel}")

    print("\n--- KONEČNÝ STAV DAT V PAMĚTI AGENTA ---")
    for record in agent._get_all_records():
        print(f" > Soubor: {os.path.basename(record['_source_file'])}, Záznam: {record['Jméno']} {record['Příjmení']}")
